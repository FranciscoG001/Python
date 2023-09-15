import pygame
import random
import numpy as np
from keras.models import Sequential
from keras.layers import Dense
import openpyxl
import tensorflow as tf
import contextlib
import sys
import os
# Redirecione a saída padrão para um arquivo de log vazio
f = open(os.devnull, 'w')
sys.stdout = f


"""def addValuesToExcel(valuesToExcel):
    # Get next row
    row = sheet.max_row + 1
    # Add data do Excel
    sheet[f"A{row}"] = valuesToExcel[0]
    sheet[f"B{row}"] = valuesToExcel[1]
    sheet[f"C{row}"] = valuesToExcel[2]
    sheet[f"D{row}"] = valuesToExcel[3]
    sheet[f"E{row}"] = valuesToExcel[4]
    sheet[f"F{row}"] = valuesToExcel[5]
    sheet[f"G{row}"] = valuesToExcel[6]
    sheet[f"H{row}"] = valuesToExcel[7]
    sheet[f"I{row}"] = valuesToExcel[8]
    sheet[f"J{row}"] = valuesToExcel[9]
    sheet[f"K{row}"] = valuesToExcel[10]
    sheet[f"L{row}"] = valuesToExcel[11]
    sheet[f"M{row}"] = valuesToExcel[12]
    sheet[f"N{row}"] = valuesToExcel[13]
    sheet[f"O{row}"] = valuesToExcel[14]"""

model = Sequential()
model.add(Dense(16, input_dim=14, activation='relu'))  # Ajuste o número de entradas conforme necessário
model.add(Dense(1, activation='sigmoid'))  # Saída binária para pular ou não

# Compile o modelo
model.compile(loss='binary_crossentropy', optimizer='adam')

workbook = openpyxl.load_workbook('FlappyBirdData.xlsx')
sheet = workbook.active

training_data = []
labels = []

for row in sheet.iter_rows(min_row=2, values_only=True):  # Comece da segunda linha para evitar o cabeçalho
    tick, lost, centerXBird, centeryBird, widthBird, heightBird, valLeftXFirstPipe, valRightXFirstPipe, valY_pos_up_pipe, valY_pos_down_pipe, valLeftXSecondPipe, valRightXSecondPipe, valY_pos_up_pipe2, valY_pos_down_pipe2, jump = row

    # Adicione esses dados à lista de dados de treinamento
    data_point = [tick, lost, centerXBird, centeryBird, widthBird, heightBird, valLeftXFirstPipe, valRightXFirstPipe, valY_pos_up_pipe, valY_pos_down_pipe, valLeftXSecondPipe, valRightXSecondPipe, valY_pos_up_pipe2, valY_pos_down_pipe2]
    training_data.append(data_point)

    # Adicione o rótulo (ação de pular) à lista de rótulos
    labels.append(jump)

# Create New Excel
#workbook = openpyxl.Workbook()

# Select Sheet
#sheet = workbook.active

# Add Colluns
"""sheet["A1"] = "Tick"
sheet["B1"] = "Lost"
sheet["C1"] = "CenterXBird"
sheet["D1"] = "CenterYBird"
sheet["E1"] = "WidthBird"
sheet["F1"] = "HeightBird"
sheet["G1"] = "ValLeftXFirstPipe"
sheet["H1"] = "ValRightXFirstPipe"
sheet["I1"] = "ValYFirstUpPipe"
sheet["J1"] = "ValYFirstDownPipe"
sheet["K1"] = "ValLeftXSecondPipe"
sheet["L1"] = "ValRightXSecondPipe"
sheet["M1"] = "ValYSecondUpPipe"
sheet["N1"] = "ValYSecondDownPipe"
sheet["O1"] = "Jump"""

# Init setup vars
width = 640
height = 720
tick = 0
lost = False
jump = False

running = True
start = False

countPoints = 0
textColor = (255, 255, 255)

# Init possition
bird_x = (width/2)-20
bird_y = height / 2

# Movement
velBird = 0  
gravity = 0.5
impulse = -9
velocityPipeMove = 5
pipePos_x = width
pipePos_x2 = width * 2

# Start coords
upPipPos_y = -370
upPipPos_y2 = -270
downPipPos_y = 520
downPipPos_y2 = 620

# To random pipes
upPipPos_y_min = -250
upPipPos_y_max = -550

# Load Images
backgroundImage = pygame.image.load("Images/background.jpg")
upPipeImage = pygame.image.load("Images/upPipe.png")
downPipeImage = pygame.image.load("Images/downPipe.png")
flappyBirdTextImage = pygame.image.load("Images/flappyBirdText.png")
birdImage = pygame.image.load("Images/bird.png")
playButtonImage = pygame.image.load("Images/playbtn.png")

# Scale original image
backgroundResizeImage = pygame.transform.scale(backgroundImage, (width, height))
upPipeResizeImage = pygame.transform.scale(upPipeImage, (width/7, height))
downPipeRisezeImage = pygame.transform.scale(downPipeImage, (width/7, height))
upPipeResizeImage2 = pygame.transform.scale(upPipeImage, (width/7, height))
downPipeRisezeImage2 = pygame.transform.scale(downPipeImage, (width/7, height))
flappyBirdTextResizeImage = pygame.transform.scale(flappyBirdTextImage, (width, 200))
birdResizeImage = pygame.transform.scale(birdImage, (75, 75))
playButtonResizeImage = pygame.transform.scale(playButtonImage, (150, 75))

# Pygame setup
pygame.init()
screen = pygame.display.set_mode((width, height))
clock = pygame.time.Clock()
font = pygame.font.Font(None, 100)
displayText = font.render(str(countPoints), True, textColor)

# Convert to avoid the black pixeis of image
birdResizeImage = birdResizeImage.convert_alpha()
upPipeResizeImage = upPipeResizeImage.convert_alpha()
downPipeRisezeImage = downPipeRisezeImage.convert_alpha()
upPipeResizeImage2 = upPipeResizeImage2.convert_alpha()
downPipeRisezeImage2 = downPipeRisezeImage2.convert_alpha()

# Pygame Loop
while running:

    for event in pygame.event.get():
        if event.type == pygame.QUIT:
            running = False # QUIT
        elif event.type == pygame.KEYDOWN:
            if event.key == pygame.K_RETURN:
                start = True # Start game
                tick = 0
                lost = False
                jump = False
                bird_x = (width/2)-20
                bird_y = height / 2
                velBird = 0  
                gravity = 0.5
                pipePos_x = width
                pipePos_x2 = width + 400
                countPoints = 0
                displayText = font.render(str(countPoints), True, textColor)
            if event.key == pygame.K_SPACE:
                velBird = impulse # Bird Jump
                jump = True
            if event.key == pygame.K_ESCAPE:
                start = False # Stop Game
        elif event.type == pygame.KEYUP:
            if event.key == pygame.K_SPACE:
                jump = False
    
    screen.blit(backgroundResizeImage,(0,0))
    screen.blit(flappyBirdTextResizeImage,(0,0))
    screen.blit(birdResizeImage,((width/2)-20,height/2))
    screen.blit(playButtonResizeImage,((width/2)-60,(height/2)-130))

    if start:
        
        model.fit(np.array(training_data), np.array(labels), epochs=1, verbose=0)

        # Count Ticks
        tick += 1

        # Pips Move Velocity
        pipePos_x -= velocityPipeMove
        pipePos_x2 -= velocityPipeMove

        # Bird Move
        velBird += gravity
        bird_y += velBird

        # Screen clean
        screen.fill((0,0,0))
        
        # Show all imagens
        screen.blit(backgroundResizeImage,(0,0))
        screen.blit(upPipeResizeImage, (pipePos_x, upPipPos_y))
        screen.blit(downPipeRisezeImage, (pipePos_x, downPipPos_y))
        screen.blit(upPipeResizeImage2, (pipePos_x2, upPipPos_y2))
        screen.blit(downPipeRisezeImage2, (pipePos_x2, downPipPos_y2))

        if countPoints < 10:
            screen.blit(displayText,(width/2,height-600))
        else:
            screen.blit(displayText,((width/2)-30,height-600))
        screen.blit(birdResizeImage,(bird_x,bird_y))
        
        # Reset Pip
        if pipePos_x + upPipeResizeImage.get_width() < -50:
            upPipPos_y = random.randint(upPipPos_y_max,upPipPos_y_min) # New position
            downPipPos_y = upPipPos_y + 890
            pipePos_x = width
        # Reset Pip2
        if pipePos_x2 + upPipeResizeImage2.get_width() < -50:
            upPipPos_y2 = random.randint(upPipPos_y_max,upPipPos_y_min) # New position
            downPipPos_y2 = upPipPos_y2 + 890
            pipePos_x2 = width

        # Ground Touch
        if bird_y > 600:
            start = False
            lost = True
        # Sky Touch
        if bird_y < 0:
            start = False
            lost = True

        # Count Pip
        if bird_x > pipePos_x + upPipeResizeImage.get_width() and not passagem_registrada:
            countPoints += 1
            displayText = font.render(str(countPoints), True, textColor)
            passagem_registrada = True
        elif bird_x <= pipePos_x + upPipeResizeImage.get_width():
            passagem_registrada = False
        # Count Pip2
        if bird_x > pipePos_x2 + upPipeResizeImage2.get_width() and not passagem_registrada2:
            countPoints += 1
            displayText = font.render(str(countPoints), True, textColor)
            passagem_registrada2 = True
        elif bird_x <= pipePos_x2 + upPipeResizeImage2.get_width():
            passagem_registrada2 = False

        # Touch Pipes
        bird_rect = pygame.Rect(bird_x+13, bird_y+20, birdResizeImage.get_width()-23, birdResizeImage.get_height()-39)
        up_pipe_rect = pygame.Rect(pipePos_x, 0, width/7, upPipPos_y+720)
        down_pipe_rect = pygame.Rect(pipePos_x, upPipPos_y + 890, width/7, height)
        up_pipe_rect2 = pygame.Rect(pipePos_x2, 0, width/7, upPipPos_y2+720)
        down_pipe_rect2 = pygame.Rect(pipePos_x2, upPipPos_y2 + 890, width/7, height)

        # Pipes 1
        if bird_rect.colliderect(up_pipe_rect) or bird_rect.colliderect(down_pipe_rect):
            start = False
            lost = True
        # Pipes 2
        if bird_rect.colliderect(up_pipe_rect2) or bird_rect.colliderect(down_pipe_rect2):
            start = False
            lost = True
            
        # First pipes
        valLeftXFirstPipe = up_pipe_rect.bottomleft[0]
        valRightXFirstPipe = up_pipe_rect.bottomright[0]

        valY_pos_up_pipe = up_pipe_rect.bottomleft[1]
        valY_pos_down_pipe = down_pipe_rect.topleft[1]

        # Second pipes
        valLeftXSecondPipe = up_pipe_rect2.bottomleft[0]
        valRightXSecondPipe = up_pipe_rect2.bottomright[0]

        valY_pos_up_pipe2 = up_pipe_rect2.bottomleft[1]
        valY_pos_down_pipe2 = down_pipe_rect2.topleft[1]

        #Bird Data
        widthBird = bird_rect.right - bird_rect.left
        heightBird = bird_rect.bottom - bird_rect.top
        centerXBird = bird_rect.centerx
        centeryBird = bird_rect.centery
        
        # Array with all necessary data to Machine Learning
        valuesToExcel = [tick,lost,centerXBird,centeryBird,widthBird,heightBird,valLeftXFirstPipe, valRightXFirstPipe, valY_pos_up_pipe,valY_pos_down_pipe, valLeftXSecondPipe, valRightXSecondPipe,valY_pos_up_pipe2, valY_pos_down_pipe2]
        
        # Use o modelo para tomar decisões
        prediction = model.predict(np.array([valuesToExcel]))

        if prediction > 0.5:
            # Fazer o pássaro pular
            velBird = impulse
        #addValuesToExcel(valuesToExcel)
        
    pygame.display.flip() # Update Screen
    clock.tick(60) # FPS to 60

# Save data and close the excel file
#workbook.save("FlappyBirdData.xlsx")
#workbook.close()

pygame.quit()